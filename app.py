import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

st.set_page_config(page_title="Data Summary Generator", layout="wide")

st.title("📊 Data Summary Generator")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file:

    df = pd.read_excel(uploaded_file, engine="openpyxl")
    df.columns = df.columns.str.strip()

    df['Download Date'] = pd.to_datetime(df['Download Date'], errors='coerce')
    df = df.dropna(subset=['Download Date'])

    df = df.sort_values(by=['Download Date'])

    summary_data = []

    unique_dates = df['Download Date'].drop_duplicates().reset_index(drop=True)

    for i in range(len(unique_dates)):

        current_date = unique_dates[i]
        subset = df[df['Download Date'] == current_date]

        # ✅ Detect "No Report Received"
        no_report_flag = subset.astype(str).apply(
            lambda col: col.str.contains("No Report Received", case=False, na=False)
        ).any().any()

        source = str(subset['Source'].iloc[0]).upper()

        # ✅ Date Logic
        if source == "MHRA":
            if i == 0:
                if current_date.day_name() == 'Monday':
                    from_date = current_date - pd.Timedelta(days=3)
                    to_date = current_date - pd.Timedelta(days=1)
                else:
                    from_date = current_date - pd.Timedelta(days=1)
                    to_date = current_date - pd.Timedelta(days=1)
            else:
                prev_date = unique_dates[i - 1]
                from_date = prev_date
                to_date = current_date - pd.Timedelta(days=1)

        elif source in ["ADIS", "NA"]:
            weekday = current_date.weekday()
            from_date = current_date - pd.Timedelta(days=weekday)
            to_date = current_date

        else:
            from_date = None
            to_date = None

        if no_report_flag:
            summary_data.append({
                'Downloaded Date': current_date,
                'Source': "",
                'From': "",
                'To': "",
                'Total Number': "",
                'Valid': "",
                'Non-Valid': "",
                'No_Report': True
            })
        else:
            summary_data.append({
                'Downloaded Date': current_date,
                'Source': source,
                'From': from_date,
                'To': to_date,
                'Total Number': len(subset),
                'Valid': (subset['Validity'] == 'Valid').sum(),
                'Non-Valid': (subset['Validity'] == 'Non-Valid').sum(),
                'No_Report': False
            })

    summary = pd.DataFrame(summary_data)

    # ✅ Format dates (only valid ones)
    for col in ['Downloaded Date', 'From', 'To']:
        summary[col] = pd.to_datetime(summary[col], errors='coerce').dt.strftime('%d-%b-%y')

    st.dataframe(summary.drop(columns=['No_Report']), use_container_width=True)

    # ✅ ✅ CREATE EXCEL WITH MERGED CELLS
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    headers = ['Downloaded Date', 'Source', 'From', 'To', 'Total Number', 'Valid', 'Non-Valid']
    ws.append(headers)

    for idx, row in summary.iterrows():
        excel_row = idx + 2

        ws.cell(row=excel_row, column=1, value=row['Downloaded Date'])

        if row['No_Report']:
            # ✅ Merge columns B to G
            ws.merge_cells(start_row=excel_row, start_column=2, end_row=excel_row, end_column=7)
            cell = ws.cell(row=excel_row, column=2)
            cell.value = "No Report Received"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        else:
            ws.cell(row=excel_row, column=2, value=row['Source'])
            ws.cell(row=excel_row, column=3, value=row['From'])
            ws.cell(row=excel_row, column=4, value=row['To'])
            ws.cell(row=excel_row, column=5, value=row['Total Number'])
            ws.cell(row=excel_row, column=6, value=row['Valid'])
            ws.cell(row=excel_row, column=7, value=row['Non-Valid'])

    # ✅ Save Excel
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.download_button(
        "⬇️ Download Excel (Merged Format)",
        data=output,
        file_name="Final_Merged_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
